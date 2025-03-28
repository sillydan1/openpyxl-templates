from datetime import datetime

from future.utils import with_metaclass
from openpyxl import Workbook, load_workbook

from openpyxl_templates.exceptions import OpenpyxlTemplateException
from openpyxl_templates.styles import DefaultStyleSet, StyleSet
from openpyxl_templates.templated_sheet import TemplatedWorksheet
from openpyxl_templates.utils import OrderedType, Typed


class SheetnamesNotUnique(OpenpyxlTemplateException):
    def __init__(self, templated_workbook):
        super(SheetnamesNotUnique, self).__init__(
            "Sheetnames are not unique on TemplatedWorkbook '%s'." % type(templated_workbook).__name__)


class MultipleActiveSheets(OpenpyxlTemplateException):
    def __init__(self, templated_workbook):
        super(MultipleActiveSheets, self).__init__(
            "The TemplatedWorkbook '%s' has multiple active sheets." % type(templated_workbook).__name__)


class TemplatedWorkbook(with_metaclass(OrderedType)):  # type: ignore
    item_class = TemplatedWorksheet
    template_styles: StyleSet
    timestamp = Typed("timestamp", expected_types=(str, bool), value=False)
    _default_timestamp: str = "%Y%m%d_%H%M%S"
    _file_extension: str = "xlsx"
    workbook: Workbook

    def __init__(self, file=None, template_styles=None, timestamp=None, templated_sheets=None, keep_vba=False, data_only=False, keep_links=True):
        super(TemplatedWorkbook, self).__init__()
        self.templated_sheets: list = []
        self.workbook = load_workbook(
            filename=file,
            data_only=data_only,
            keep_vba=keep_vba,
            keep_links=keep_links
        ) if file else Workbook()
        self.template_styles = template_styles or DefaultStyleSet()
        self.timestamp = timestamp or False
        for sheetname, templated_sheet in self._items.items():
            self.add_templated_sheet(templated_sheet, sheetname=sheetname, add_to_self=False)
        for templated_sheet in templated_sheets or []:
            self.add_templated_sheet(sheet=templated_sheet, sheetname=templated_sheet.sheetname, add_to_self=True)
        self._validate()

    def _validate(self):
        self._check_unique_sheetnames()
        self._check_only_one_active()

    def _check_unique_sheetnames(self):
        if len(set(templated_sheet.sheetname for templated_sheet in self.templated_sheets)) < len(self.templated_sheets):
            raise SheetnamesNotUnique(self)

    def _check_only_one_active(self):
        if len(tuple(sheet for sheet in self.templated_sheets if sheet.active)) > 1:
            raise MultipleActiveSheets(self)

    def add_templated_sheet(self, sheet, sheetname=None, add_to_self=True):
        if sheetname and not sheet._sheetname:
            sheet._sheetname = sheetname
        sheet.workbook = self.workbook
        sheet.template_styles = self.template_styles
        self.templated_sheets.append(sheet)
        return sheet
        # TODO: Parse sheetname to an attribute? Or removing add to self all together?
        # if add_to_self:
        #     setattr(self, sheet.sheetname, sheet)

    def remove_all_sheets(self):
        for sheetname in self.workbook.sheetnames:
            del self.workbook[sheetname]

    def save(self, filename):
        if self.timestamp:
            filename = self.timestamp_filename(filename)
        self.sort_worksheets()
        self.workbook.save(filename)
        return filename

    def sort_worksheets(self):
        order = {}
        index = 0
        active_index = 0
        for templated_sheet in self.templated_sheets:
            order[templated_sheet.sheetname] = index
            if templated_sheet.active:
                active_index = index
            index += 1
        for sheetname in self.workbook.sheetnames:
            if sheetname not in order:
                order[sheetname] = index
                index += 1
        # TODO: fiddling with a private member is a big nono, but this is good enough for now
        self.workbook._sheets = sorted(self.workbook._sheets, key=lambda s: order[s.title])  # type: ignore
        self.workbook.active = active_index

    def timestamp_filename(self, filename):
        return "%s_%s.%s" % (
            filename.strip(".%s" % self._file_extension),
            datetime.now().strftime(
                self.timestamp
                if isinstance(self.timestamp, str)
                else self._default_timestamp
            ),
            self._file_extension)

    @property
    def sheetnames(self):
        return self.workbook.sheetnames

    def create_sheet(self, title=None, index=None):
        return self.workbook.create_sheet(title, index)
